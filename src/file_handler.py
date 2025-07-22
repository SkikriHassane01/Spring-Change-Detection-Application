"""
This script will validate the excel file uploaded by the user then validate the crucial columns
after that creating the excel output when the user press on download button
"""
import streamlit as st
#__TODO: import libraries_______________________________________________
import io
from typing import Any, Tuple, Optional

import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

from config import UPLOAD_CONFIG, REQUIRED_COLUMNS

class FileHandler:
    """Handles validation and export of Excel files."""

    #__TODO: Validate the uploaded excel buffer_______________________________________________
    @staticmethod
    def validate_excel_file(
        file: Any, file_label: str
    ) -> Tuple[bool, str, Optional[pd.DataFrame]]:
        """
        Args:
            file: Uploaded file.
            file_label: A label for the file (e.g., "old", "new").

        Returns:
            Tuple containing:
              - validity (bool)
              - message (str)
              - DataFrame if valid, else None
        """
        if not file:
            return False, f"No '{file_label}' file uploaded.", None

        try:
            df = (
                pd.read_excel(
                    file,
                    engine="openpyxl",
                    sheet_name=UPLOAD_CONFIG["sheet_name"],
                    skiprows=UPLOAD_CONFIG["skip_rows"],
                )
                .reset_index(drop=True)
            )
        except Exception as e:
            return False, f"Error reading '{file_label}' file: {e}", None

        if df.empty:
            return False, f"'{file_label}' file is empty.", None

        is_valid, msg = FileHandler._validate_columns(df)
        if not is_valid:
            return False, msg, None

        return True, "File uploaded successfully.", df
    
    #__TODO: Validate the crucial columns_______________________________________________
    @staticmethod
    def _validate_columns(df: pd.DataFrame) -> Tuple[bool, str]:
        """
        Check for required columns

        Args:
            df: DataFrame to validate.

        Returns:
            validity and error message if invalid.
        """
        required_cols = [
            REQUIRED_COLUMNS["mass"],
            REQUIRED_COLUMNS["reference"],
        ]
        missing = [col for col in required_cols if col not in df.columns]
        if missing:
            return False, f"Missing columns: {', '.join(missing)}."
        return True, ""
    
    #__TODO: Create the excel output _______________________________________________
    @staticmethod
    def create_excel_bytes(
        results: pd.DataFrame
    ) -> bytes:
        """
        Generate a styled Excel report as bytes.

        Args:
            results: Analysis results with metadata.

        Returns:
            Byte content of the Excel file.
        """
        
        new_df = st.session_state.get('input_excel_new')
        if results is None or new_df is None:
            raise ValueError("both 'results' and 'new df' dataframes are required.")

        export_df = new_df.copy()
        
        metadata_cols = [
            'Old Reference', 'New Reference',
            'Mass Status', 'Change Type',
            'Cell ID New', 'Cell ID Old'
        ]
        
        for col in metadata_cols:
            export_df[col] = results[col]

        # ensure ascending order by new-cell ID
        export_df = export_df.sort_values('Cell ID New', ascending=True).reset_index(drop=True)
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            export_df.to_excel(writer, sheet_name='PTA Analysis', index=False)
            wb = writer.book
            ws = writer.sheets['PTA Analysis']

            # styling definitions
            header_fill = PatternFill('solid', fgColor='4F81BD')
            header_font = Font(bold=True, color='FFFFFF', size=12)
            header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
            thin = Border(*(Side('thin'),) * 4)
            fill_new = PatternFill('solid', fgColor='FF5733')
            fill_spring = PatternFill('solid', fgColor='B4C6E7')

            # style header row
            for col_idx in range(1, export_df.shape[1] + 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_align
                cell.border = thin

            max_row = export_df.shape[0] + 1
            max_col = export_df.shape[1]
            change_idx = export_df.columns.get_loc('Change Type') + 1

            # apply row & cell styling
            for r in range(2, max_row + 1):
                typ = export_df.at[r - 2, 'Change Type']
                if typ == 'New':
                    for c in range(1, max_col + 1): ws.cell(r, c).fill = fill_new
                elif typ == 'Spring Changed':
                    for c in range(1, max_col + 1):
                        ws.cell(r, c).fill = fill_spring
            # auto-size columns & rows
            for col in ws.columns:
                length = max(len(str(cell.value)) for cell in col if cell.value)
                ws.column_dimensions[col[0].column_letter].width = min(length + 3, 30)
            for r in range(2, max_row + 1):
                ws.row_dimensions[r].height = 18

        output.seek(0)
        return output.getvalue()